from flask import Flask , render_template , request , redirect , url_for, flash
from openpyxl import load_workbook
from flask_bootstrap import Bootstrap
import re
from openpyxl.utils.exceptions import InvalidFileException


dictionary = {}

app = Flask (__name__)
bootstrap = Bootstrap(app)

@app.route ( '/' )
def index():
    return render_template ('index.html', dictionary=dictionary)

@app.route('/upload', methods=['GET', 'POST'])
def upload():
    try:
        # Obtener el archivo Excel
        file_excel = request.files['file_excel']

        # Leer el archivo Excel
        wb = load_workbook(file_excel)

        # Seleccionar la hoja de trabajo
        sheet = wb['Base de Datos']

        # Recorrer las filas y guardar las palabras y pronunciaciones
        for row in sheet.iter_rows(min_row=2):
            palabra = row[1].value.strip().lower() if row[2].value else ''
            pronunciation = row[2].value.strip().lower() if row[2].value else ''
            dictionary[palabra] = pronunciation

        # Redirigir a la página principal
        return redirect(url_for('index'))

    except InvalidFileException:
        flash('El archivo proporcionado no es un archivo Excel válido.', 'error')
        return redirect(url_for('upload'))

    except Exception as e:
        flash(f'Ocurrió un error inesperado: {str(e)}', 'error')
        return redirect(url_for('upload'))

@app.route('/traducer', methods=['GET', 'POST'])
def traducer():
    texto = request.form['texto'].lower()
    transduction = ''
    color = ''

    if request.method == 'POST':
        texto

    # Separación por palabras y signos de puntuación
    palabras = re.findall(r'\b\w+\b|[.,!?;]', texto)

    # Traducción y concatenación
    palabras_color = []
    for palabra in palabras:
        palabra_color = {'palabra': '', 'color': 'black'}
        if palabra in dictionary:
            palabra_color['palabra'] = dictionary[palabra ]
        else:
            palabra_color['palabra'] = palabra
            if re.match(r'\b\w+\b', palabra):
                palabra_color['color'] = 'red'
        palabras_color.append(palabra_color)

    return render_template('index.html', traduccion=transduction, color=color, palabras_color=palabras_color, texto=texto, dictionary=dictionary)

@app.errorhandler(404)
def page_not_found(error: Exception):
    return render_template('404.html', error=error)

@app.errorhandler(500)
def internal_server_error():
    return render_template('500.html')


if __name__ == '__main__':
    app.run ( debug=True )
